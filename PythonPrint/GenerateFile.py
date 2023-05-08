import openpyxl
import qrcode
from PIL import Image
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

def GenerateExcel(path='./assets/example.xlsx',packid:str="Pack0001",snList:any=[{'SN':'SN000000001','Lot':'Lot1','Date':'20230101','Remark':'良品'},{'SN':'SN000000002','Lot':'Lot2','Date':'20230102','Remark':'不良品'}]):
    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()

    # 选择第一个工作表
    worksheet = workbook.active

    # 设置单元格的值
    worksheet['A1'] = ''
    worksheet['A2'] = ('出库信息%s' % packid)

    worksheet['A3'] = '产品SN'
    worksheet['B3'] = '批次'
    worksheet['C3'] = '生产日期'
    worksheet['D3'] = '备注'

    rowInt=4
    for item in snList:
        worksheet['A'+str(rowInt)] = item['SN']
        worksheet['B'+str(rowInt)] = item['Lot']
        worksheet['C'+str(rowInt)] = item['Date']
        worksheet['D'+str(rowInt)] = item['Remark']
        rowInt=rowInt+1

    # 合并单元格
    worksheet.merge_cells('A1:D1')
    worksheet.merge_cells('A2:D2')

    # 设置单元格的宽度和高度
    worksheet.column_dimensions['A'].width = 15
    worksheet.column_dimensions['B'].width = 10
    worksheet.column_dimensions['C'].width = 10
    worksheet.column_dimensions['D'].width = 10
    worksheet.row_dimensions[1].height = 5
    worksheet.row_dimensions[2].height = 41
    worksheet.row_dimensions[3].height = 30

    # 插入图片或二维码
    # 生成二维码图像
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=2, border=3)
    qr.add_data(packid)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')

    # 将图像插入单元格中
    img_file = './assets/qr_code.png'
    img.save(img_file)
    img_obj = Image(img_file)

    # 将图像插入单元格B2并设置距离
    worksheet.add_image(img_obj, 'D2')

    # 添加所有边框
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))

    # 设置单元格的对齐方式
    for col in worksheet.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
    worksheet['A1'].border = worksheet['B1'].border=worksheet['C1'].border=worksheet['D1'].border=Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'))

    worksheet['A2'].border = worksheet['B2'].border=worksheet['C2'].border=worksheet['D2'].border=Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        bottom=Side(style='thin'))

    # 保存工作簿
    workbook.save(path)